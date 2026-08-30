"""
Lectura de la planilla original (solo lectura, nunca se escribe sobre el
archivo de entrada) y escritura de la planilla enriquecida derivada.
"""

import datetime as _dt
import json
from pathlib import Path
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[2]
PLANILLA_ORIGINAL = PROJECT_ROOT / "input_spreadsheet.xlsx"


def leer_planilla_principal() -> list[dict]:
    """Lee la hoja 'Ranking actualizado' (header en la fila 4) en modo
    solo-lectura. Nunca abre el archivo para escritura."""
    wb = openpyxl.load_workbook(PLANILLA_ORIGINAL, data_only=True, read_only=True)
    ws = wb["Ranking actualizado"]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    header = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]

    registros = []
    for r in rows[1:]:
        row_dict = {header[i]: r[i] for i in range(len(header)) if i < len(r)}
        if row_dict.get("URL"):  # descarta filas vacías al final del rango
            registros.append(row_dict)

    wb.close()
    return registros


_TIPOS_EXCEL_NATIVOS = (str, int, float, bool, _dt.datetime, _dt.date, type(None))


def _valor_para_celda(valor):
    """
    openpyxl solo acepta tipos "planos" (str/int/float/bool/date/None) en
    una celda — un dict o list (ej. address_structured, address_text,
    needs_manual_review_reasons, jsonld) rompe wb.save() con un error
    críptico recién al guardar, potencialmente después de procesar las
    511 filas. Se serializan a JSON (conservando el dato completo, no se
    pierde trazabilidad) en vez de intentar aplanarlos o descartarlos.
    """
    if isinstance(valor, _TIPOS_EXCEL_NATIVOS):
        return valor
    return json.dumps(valor, ensure_ascii=False, default=str)


def escribir_planilla_enriquecida(
    hojas: dict[str, list[dict]], output_path: str
) -> None:
    """
    hojas: {"PROPIEDADES_ENRIQUECIDAS": [...], "SCRAPING_ERRORES": [...], ...}
    Escribe SIEMPRE en un archivo nuevo derivado, nunca sobre el original.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for nombre_hoja, registros in hojas.items():
        ws = wb.create_sheet(title=nombre_hoja[:31])  # Excel limita a 31 caracteres
        if not registros:
            ws.append(["(sin registros)"])
            continue

        # Unión de columnas de TODOS los registros (no solo el primero):
        # los campos varían por dominio (ej. latitude_sitio solo aparece
        # en RE/MAX), y tomar solo las claves del primer registro dejaría
        # esas columnas afuera en silencio para el resto de las filas.
        columnas: list = []
        vistas = set()
        for reg in registros:
            for c in reg.keys():
                if c not in vistas:
                    vistas.add(c)
                    columnas.append(c)

        ws.append(columnas)
        for reg in registros:
            ws.append([_valor_para_celda(reg.get(c)) for c in columnas])

    wb.save(output_path)
